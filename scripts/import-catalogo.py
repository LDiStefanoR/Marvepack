# -*- coding: utf-8 -*-
"""Importa el Excel TPV y descarga fotos representativas por familia de producto."""
from __future__ import annotations

import json
import re
import ssl
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL = Path(r"C:\Users\leord\Downloads\ProductosTPV 23-8 exel.xlsx")
IMG_DIR = ROOT / "recursos" / "imagenes-productos"
WEB_DIR = ROOT / "recursos" / "web"
OUT_JSON = ROOT / "data" / "productos.json"

# Unsplash (licencia Unsplash): fotos de stock por familia, no SKU a SKU.
IMAGES: dict[str, str] = {
    "bolsa-zipper": "https://images.pexels.com/photos/4239013/pexels-photo-4239013.jpeg?auto=compress&cs=tinysrgb&w=900",
    "bolsa-ecommerce": "https://images.unsplash.com/photo-1586528116311-ad8dd3c8310d?auto=format&fit=crop&w=900&q=80",
    "bolsa-friselina": "https://images.pexels.com/photos/1152077/pexels-photo-1152077.jpeg?auto=compress&cs=tinysrgb&w=900",
    "bolsa-kraft": "https://images.unsplash.com/photo-1607083206968-13611e3d76db?auto=format&fit=crop&w=900&q=80",
    "bolsa-rinon": "https://images.unsplash.com/photo-1513201099705-a9746e1e201f?auto=format&fit=crop&w=900&q=80",
    "bolsa-plastica": "https://images.unsplash.com/photo-1615485290382-441e4d049cb5?auto=format&fit=crop&w=900&q=80",
    "bolsa-delivery": "https://images.unsplash.com/photo-1526367790999-0150786686a2?auto=format&fit=crop&w=900&q=80",
    "caja-pizza": "https://images.unsplash.com/photo-1513104890138-7c749659a591?auto=format&fit=crop&w=900&q=80",
    "caja-comida": "https://images.unsplash.com/photo-1565299624946-b28f40a0ae38?auto=format&fit=crop&w=900&q=80",
    "caja-regalo": "https://images.unsplash.com/photo-1549465220-1a8b9238cd48?auto=format&fit=crop&w=900&q=80",
    "globos": "https://images.unsplash.com/photo-1530103862676-de8c9debad1d?auto=format&fit=crop&w=900&q=80",
    "cotillon": "https://images.pexels.com/photos/1303081/pexels-photo-1303081.jpeg?auto=compress&cs=tinysrgb&w=900",
    "film": "https://images.unsplash.com/photo-1583947215259-38e31be8751f?auto=format&fit=crop&w=900&q=80",
    "burbuja": "https://images.unsplash.com/photo-1607344645866-009c320b63e0?auto=format&fit=crop&w=900&q=80",
    "cinta": "https://images.unsplash.com/photo-1558618666-fcd25c85cd64?auto=format&fit=crop&w=900&q=80",
    "bandeja-expandido": "https://images.unsplash.com/photo-1607623814075-e51df1bdc82f?auto=format&fit=crop&w=900&q=80",
    "pote": "https://images.unsplash.com/photo-1584473457406-6240486418e9?auto=format&fit=crop&w=900&q=80",
    "bandeja": "https://images.unsplash.com/photo-1556910103-1c02745aae4d?auto=format&fit=crop&w=900&q=80",
    "vaso": "https://images.unsplash.com/photo-1572490122747-3968b75cc699?auto=format&fit=crop&w=900&q=80",
    "marmita": "https://images.unsplash.com/photo-1484980972926-edee96e0960d?auto=format&fit=crop&w=900&q=80",
    "guantes": "https://images.pexels.com/photos/3985163/pexels-photo-3985163.jpeg?auto=compress&cs=tinysrgb&w=900",
    "higiene": "https://images.unsplash.com/photo-1584622650111-993a426fbf0a?auto=format&fit=crop&w=900&q=80",
    "bolsa-residuos": "https://images.unsplash.com/photo-1611284446314-60a58ac0deb9?auto=format&fit=crop&w=900&q=80",
    "monos": "https://images.unsplash.com/photo-1513885535751-8b9238bd345a?auto=format&fit=crop&w=900&q=80",
    "palitos": "https://images.unsplash.com/photo-1497032628192-86f99bcd76bc?auto=format&fit=crop&w=900&q=80",
    "reposteria": "https://images.unsplash.com/photo-1578985545062-69928b1d9587?auto=format&fit=crop&w=900&q=80",
    "rollos": "https://images.unsplash.com/photo-1599599810769-bcde5a160d32?auto=format&fit=crop&w=900&q=80",
    "libreria": "https://images.unsplash.com/photo-1456735190827-d1262f71b8a3?auto=format&fit=crop&w=900&q=80",
    "platos": "https://images.unsplash.com/photo-1574484284002-952d92456975?auto=format&fit=crop&w=900&q=80",
    "cubiertos": "https://images.unsplash.com/photo-1556911220-bff31c812dba?auto=format&fit=crop&w=900&q=80",
    "default": "https://images.unsplash.com/photo-1604719312566-8912e9227c6a?auto=format&fit=crop&w=900&q=80",
}

WEB_IMAGES = {
    "local-mostrador.jpg": "https://images.unsplash.com/photo-1441986300917-64674bd600d8?auto=format&fit=crop&w=1200&q=80",
}

SECTION_BLURB = {
    "BOLSAS": "Bolsas para delivery, mostrador y e-commerce. Consultanos medidas y cantidades.",
    "CAJAS MICRO": "Cajas microcorrugadas para hamburguesas, pizzas y delivery caliente.",
    "CAJAS Y ESTUCHES": "Cajas y estuches para regalos, desayunos y pastelería.",
    "COTILLON": "Cotillón, globos y detalles para fiestas. Stock para reponer al toque.",
    "DESCARTABLES": "Descartables de uso diario para gastronomía y comercio.",
    "EMBALAJE": "Cintas, film stretch y burbuja para armar y proteger envíos.",
    "EXPANDIDO": "Bandejas y obleas de expandido para carnicería y mostrador.",
    "HIGIENE": "Higiene y limpieza para el local: rollos, guantes y bolsas.",
    "LIBRERIA": "Librería y accesorios: cintas, moños, palitos y más.",
    "PERSONALIZADOS": "Bolsas personalizadas. Te asesoramos en medidas y cantidades.",
    "POTES": "Potes con tapa para salsas, postres y take away.",
    "REPOSTERIA": "Insumos y descartables de repostería para el obrador.",
    "ROLLOS": "Rollos de manteca, aluminio, sulfito y envasado.",
    "TERMICOS": "Vasos y marmitas térmicas para llevar caliente.",
    "VARIOS": "Bandejas y descartables varios para el día a día.",
}

SECTION_DEFAULT = {
    "BOLSAS": "bolsa-plastica",
    "CAJAS MICRO": "caja-comida",
    "CAJAS Y ESTUCHES": "caja-regalo",
    "COTILLON": "cotillon",
    "DESCARTABLES": "default",
    "EMBALAJE": "cinta",
    "EXPANDIDO": "bandeja-expandido",
    "HIGIENE": "higiene",
    "LIBRERIA": "libreria",
    "PERSONALIZADOS": "bolsa-friselina",
    "POTES": "pote",
    "REPOSTERIA": "reposteria",
    "ROLLOS": "rollos",
    "TERMICOS": "vaso",
    "VARIOS": "bandeja",
}

RULES: list[tuple[list[str], str]] = [
    (["zipper", "cierre"], "bolsa-zipper"),
    (["ecommerce", "e commerce", "e-commerce"], "bolsa-ecommerce"),
    (["friselina"], "bolsa-friselina"),
    (["kraft"], "bolsa-kraft"),
    (["riñon", "rinon", "riño", "rino"], "bolsa-rinon"),
    (["polipropileno", "cristal", "arranque"], "bolsa-plastica"),
    (["manija", "delivery"], "bolsa-delivery"),
    (["pizza"], "caja-pizza"),
    (["hamburg", "torpedo", "raviole", "caja micro", "caja delivery"], "caja-comida"),
    (["drip cake", "cupcake", "box", "estuche", "bombonera", "corazon", "perfume"], "caja-regalo"),
    (["globo"], "globos"),
    (["lente", "antifaz", "pulsera", "anillo", "luz", "fluor", "neon", "shimmer"], "cotillon"),
    (["stretch", "strech", "film", "pvc"], "film"),
    (["burbuja"], "burbuja"),
    (["cinta"], "cinta"),
    (["expandido", "oblea"], "bandeja-expandido"),
    (["pote", "bisagra pet"], "pote"),
    (["bandeja"], "bandeja"),
    (["vaso"], "vaso"),
    (["marmita"], "marmita"),
    (["guante"], "guantes"),
    (["higienico", "higenico", "cocina", "pañu", "panu", "bobina limpieza", "snif"], "higiene"),
    (["residuo", "consorcio", "concorcio"], "bolsa-residuos"),
    (["moño", "mono"], "monos"),
    (["palito"], "palitos"),
    (["pirotin", "molde", "colorante", "dulce", "torta", "alfajor", "plato aluminio"], "reposteria"),
    (["aluminio", "manteca", "sulfito", "vacio", "vacío", "corrugado", "envasar"], "rollos"),
    (["silicona", "abroch", "resma"], "libreria"),
    (["plato"], "platos"),
    (["cubierto", "tenedor", "cuchillo", "cuchar"], "cubiertos"),
]


def norm(text: str) -> str:
    return (
        text.lower()
        .replace("ñ", "n")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )


def image_key(nombre: str, seccion: str) -> str:
    n = norm(nombre)
    for keys, slug in RULES:
        if any(k in n for k in keys):
            return slug
    return SECTION_DEFAULT.get(seccion, "default")


def download(url: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 4000:
        return
    ctx = ssl.create_default_context()
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 MarvePackCatalog/1.0"},
    )
    with urllib.request.urlopen(req, context=ctx, timeout=40) as res:
        dest.write_bytes(res.read())
    print("ok", dest.name, dest.stat().st_size)


def main() -> None:
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb["ProductosTPV"]
    products: list[dict] = []
    seen: dict[str, int] = {}

    for row in ws.iter_rows(min_row=11, values_only=True):
        codigo, nombre = row[0], row[2]
        precio, mayorista, seccion = row[7], row[8], row[9]
        if not nombre:
            continue
        nombre_s = str(nombre).strip()
        seccion_s = str(seccion).strip() if seccion else "VARIOS"
        codigo_s = re.sub(r"\s+", "", str(codigo).strip()) if codigo is not None else "s"
        seen[codigo_s] = seen.get(codigo_s, 0) + 1
        pid = f"mp-{codigo_s}" if seen[codigo_s] == 1 else f"mp-{codigo_s}-{seen[codigo_s]}"
        slug = image_key(nombre_s, seccion_s)
        precio_n = float(precio) if precio not in (None, "") else 0
        mayor_n = float(mayorista) if mayorista not in (None, "") else None
        item = {
            "id": pid,
            "codigo": str(codigo).strip() if codigo is not None else "",
            "nombre": nombre_s,
            "seccion": seccion_s,
            "precio": round(precio_n),
            "imagen": f"/recursos/imagenes-productos/{slug}.jpg",
            "descripcion": SECTION_BLURB.get(seccion_s, "Consultá stock y precio por WhatsApp."),
        }
        if mayor_n is not None:
            item["precioMayorista"] = round(mayor_n)
        products.append(item)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    print("productos", len(products), "->", OUT_JSON)

    for slug, url in IMAGES.items():
        try:
            download(url, IMG_DIR / f"{slug}.jpg")
        except Exception as exc:
            print("FAIL", slug, exc)

    for name, url in WEB_IMAGES.items():
        try:
            download(url, WEB_DIR / name)
        except Exception as exc:
            print("FAIL", name, exc)


if __name__ == "__main__":
    main()
